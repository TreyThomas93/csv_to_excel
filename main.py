import csv
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from win32com.client import Dispatch
import time
import os
from pprint import pprint
from playsound import playsound

class CSVToExcel():

    def __init__(self, csv_filename, new_csv_filename, excel_filename, sheetname, range):

        self.csv_filename = csv_filename

        self.excel_filename = excel_filename

        self.new_csv_filename = new_csv_filename

        self.sheetname = sheetname

        self.range = range

        self.book = load_workbook(excel_filename)

    def convertCSV(self):

        lst = []

        with open(self.csv_filename, 'r') as csvfile:

            csvreader = csv.reader(csvfile)

            fields = next(csvreader)

            for row in csvreader:

                try:

                    convert = row[0].replace(";", ",")

                    splt = convert.split(",")[1:-1]

                    splt[4] = datetime.strptime(
                        splt[4], "%m/%d/%y %I:%M %p").strftime("%m/%d/%y %H:%M")

                    lst.append(splt)

                except:

                    pass
        
        with open(self.new_csv_filename, 'w', newline='') as myfile:

            writer = csv.DictWriter(myfile, fieldnames=[
                "Strategy", "Side", "Amount", "Price", "Date/Time", "Trade P/L", "P/L", "Position"])

            writer.writeheader()

            writer = csv.writer(myfile)
            
            for i in lst:

                writer.writerow(i)
        
    def clearCurrentCells(self):

        sheet = self.book[self.sheetname]

        for row in sheet[self.range]:
            for cell in row:
                cell.value = None

        self.book.save(self.excel_filename)

    def fixDuplicates(self, df):

        rows_to_delete = []

        for index, row in df.iterrows():

            try:

                current_side = df.loc[index, "Side"]

                next_side = df.loc[index + 1, "Side"]

                third_side = df.loc[index + 2, "Side"]

                current_price = df.loc[index, "Price"]

                next_price = df.loc[index + 1, "Price"]

                third_price = df.loc[index + 2, "Price"]

                current_date = df.loc[index, "Date/Time"]

                next_date = df.loc[index + 1, "Date/Time"]

                third_date = df.loc[index + 2, "Date/Time"]

                sell = "Sell to Close"

                buy = "Buy to Open"

                if current_side == buy and next_side == buy or current_side == sell and next_side == sell:

                    if current_price == third_price:

                        # REMOVE CURRENT INDEX
                        rows_to_delete.append(index)

                    elif next_price == third_price:

                        # REMOVE LAST INDEX
                        rows_to_delete.append(index + 1)

                    else:

                        if third_price == df.loc[index + 3, "Price"]:

                            # REMOVE THIRD INDEX
                            rows_to_delete.append(index + 2)

                        else:

                            # REMOVE LAST INDEX
                            rows_to_delete.append(index + 1)

            except:

                pass

        fixed = df.drop(rows_to_delete)

        fixed = fixed.reset_index(drop=True)

        return fixed

    def addCSVToExcel(self):

        df = pd.read_csv(self.new_csv_filename)

        df = self.fixDuplicates(df)

        with pd.ExcelWriter(self.excel_filename, engine='openpyxl') as writer:

            writer.book = self.book

            writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)

            df.to_excel(writer, sheet_name=self.sheetname, startrow=0,
                        startcol=0, engine='openpyxl', merge_cells=False)

            self.book.save(self.excel_filename)


if __name__ == "__main__":

    # REQUIRED #############################################################
    excel_filename = "Book1.xlsx"  # NAME OF WHATEVER EXCEL FILE YOU ARE USING
    
    sheetname = "TOS SHEET ONE"  # NAME OF WHATEVER SHEET IN EXCEL FILE YOU WANT THE CSV DATA TO BE ADDED TO, MUST BE EXACT
    
    cell_range = "A1:I500"  # RANGE OF CELLS TO REMOVE AND ADD DATA TO
    ########################################################################

    csv_filename = "StrategyReports.csv"

    new_csv_filename = "StrategyReportsConverted.csv"

    csv_to_excel = CSVToExcel(csv_filename, new_csv_filename, excel_filename, sheetname, cell_range)

    os.system("cls")

    print("LISTENING FOR CSV...")

    exe_running = False

    while True:

        try:

            for file in os.listdir(os.getcwd()):

                if file.endswith(".csv"):

                    if file != new_csv_filename:

                        new_file_name = file.split('_')

                        if len(new_file_name) > 1:

                            # CHANGE CSV FILENAME
                            os.rename(file, f"{new_file_name[0].strip()}.csv")
            
            path = csv_filename

            new_path = f"{os.getcwd()}/{new_csv_filename}"

            if os.path.exists(path):

                print("\nCSV FOUND...\n")

                time.sleep(1)

                if exe_running:

                    print("KILLING EXISTING EXCEL EXE...\n")
                    # KILL OPEN EXCEL EXE
                    os.system("taskkill /f /im excel.exe")
                    print("\n")

                time.sleep(1)

                print("CONVERTING CSV, CLEARING CURRENT CELLS, ADDING NEW CSV DATA TO EXCEL...\n")
                # RUN 
                csv_to_excel.convertCSV()

                csv_to_excel.clearCurrentCells()

                csv_to_excel.addCSVToExcel()

                print("STARTING EXCEL EXE...\n")
                # START EXCEL EXE
                os.system(f"start EXCEL.EXE {excel_filename}")

                exe_running = True

                time.sleep(3)

                print("REMOVING EXISTING CSV...\n")
                # REMOVE CSV FILE
                os.remove(path)
                print("CSV FILE REMOVED...\n")

                os.system("cls")

                print("LISTENING FOR CSV...")

        except Exception as e:

            for _ in range(3):

                playsound('error.mp3')

            raise Exception(f"ERROR - {e}")

        time.sleep(1)